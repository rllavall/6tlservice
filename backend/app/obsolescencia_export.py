"""Exporta el report de obsolescencia de un banco (dict de obsolescencia_banco)
a Excel (.xlsx) y PDF. No toca BD."""
from __future__ import annotations

import io

# (clave en el dict de informe_banco, encabezado visible)
_COLUMNAS = [
    ("posicion", "Posición"),
    ("part_number", "P/N 6TL"),
    ("fabricante", "Fabricante"),
    ("pn_fabricante", "P/N fabricante"),
    ("descripcion", "Descripción"),
    ("numero_serie", "Nº serie"),
    ("estado_ciclo_vida", "Estado"),
    ("ciclo_vida_fecha", "Fecha evento"),
    ("ciclo_vida_verificado_en", "Verificado"),
    ("ciclo_vida_url", "Fuente"),
    ("ciclo_vida_resumen", "Resumen"),
    ("ciclo_vida_cita", "Cita"),
]
_CAB = dict(_COLUMNAS)


def _txt(v) -> str:
    return "" if v is None else str(v)


def _subtitulo(b: dict, r: dict) -> str:
    return (f"Cliente: {_txt(b['cliente'])} · Estado: {b['estado']} · "
            f"En riesgo: {r['en_riesgo']}/{r['total']} · Sin verificar: {r['sin_verificar']}")


def a_xlsx(informe: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    b, r = informe["banco"], informe["resumen"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Componentes"
    ws.append([f"Obsolescencia banco {b['numero_serie']} ({b['producto']})"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([_subtitulo(b, r)])
    ws.append([])
    ws.append([cab for _, cab in _COLUMNAS])
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)

    riesgo_fill = PatternFill("solid", fgColor="FFC7CE")
    col_estado = [c for c, _ in _COLUMNAS].index("estado_ciclo_vida") + 1
    for fila in informe["componentes"]:
        ws.append([_txt(fila.get(clave)) for clave, _ in _COLUMNAS])
        if fila["severidad"] > 0:
            ws.cell(row=ws.max_row, column=col_estado).fill = riesgo_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def a_pdf(informe: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                     Table, TableStyle)

    b, r = informe["banco"], informe["resumen"]
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Obsolescencia banco {b['numero_serie']} ({b['producto']})", estilos["Title"]),
        Paragraph(_subtitulo(b, r), estilos["Normal"]),
        Spacer(1, 6 * mm),
    ]

    columnas = ["posicion", "part_number", "fabricante", "pn_fabricante",
                "numero_serie", "estado_ciclo_vida", "ciclo_vida_fecha",
                "ciclo_vida_verificado_en"]
    datos = [[_CAB[c] for c in columnas]]
    estilo = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9e007e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])
    for i, fila in enumerate(informe["componentes"], start=1):
        datos.append([_txt(fila[c]) for c in columnas])
        if fila["severidad"] > 0:
            estilo.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFC7CE"))

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(estilo)
    elementos.append(tabla)

    buf = io.BytesIO()
    SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10 * mm,
                      rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm).build(elementos)
    return buf.getvalue()
